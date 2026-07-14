# -*- coding: utf-8 -*-
"""분석 결과 → 엑셀 리포트 (요약 + 키워드별 상세)."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "맑은 고딕"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BASE_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, size=10, bold=True)
OUR_FILL = PatternFill("solid", fgColor="FFF2CC")     # 우리 스토어 행 강조
WARN_FILL = PatternFill("solid", fgColor="FCE4E4")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

VERDICT_FILL = {
    "적정": GOOD_FILL,
    "고가 주의": WARN_FILL,
    "다소 높음": PatternFill("solid", fgColor="FFF0E0"),
    "저가 (마진 점검)": PatternFill("solid", fgColor="E0ECFF"),
}


QTY_RE = re.compile(r'(\d{1,5})\s*(?:개입|입|매입|매|켤레|쌍|조|EA|ea|pcs|PCS|P\b)')


def parse_qty(title: str):
    """상품명에서 수량 추정 (예: '10개입' → 10). 못 찾으면 None."""
    m = QTY_RE.search(title or "")
    if m:
        q = int(m.group(1))
        if 2 <= q <= 10000:
            return q
    return None


def mall_display(name: str) -> str:
    return "네이버 가격비교(카탈로그)" if name == "네이버" else name


def _sheet_name(s: str) -> str:
    s = re.sub(r'[\\/*?:\[\]]', "", s)
    return s[:31] or "sheet"


def _write_header(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
        if widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]


def build_excel(results, summary, store_name, run_date, out_path):
    wb = Workbook()

    # ---------- 요약 시트 ----------
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    ws["B2"] = f"네이버쇼핑 순위·가격 모니터링 — {store_name}"
    ws["B2"].font = Font(name=FONT, size=14, bold=True)
    ws["B3"] = f"수집 시각: {run_date}   |   기준: 네이버쇼핑 검색 정확도순(광고 제외), 가격 비교는 키워드별 상위권 상품"
    ws["B3"].font = Font(name=FONT, size=9, color="808080")

    cards = [
        ("추적 키워드", summary["keywords_total"]),
        ("노출(1000위 내)", summary["exposed"]),
        ("1페이지(10위 내)", summary["top10"]),
        ("40위 내", summary["top40"]),
        ("평균 순위", summary["avg_rank"] if summary["avg_rank"] else "-"),
        ("시장평균 대비 가격", f'{summary["avg_vs_market_pct"]:+.1f}%'
         if summary["avg_vs_market_pct"] is not None else "-"),
    ]
    r = 5
    for i, (label, val) in enumerate(cards):
        c = 2 + i * 2
        ws.cell(row=r, column=c, value=label).font = Font(name=FONT, size=9, color="595959")
        v = ws.cell(row=r + 1, column=c, value=val)
        v.font = Font(name=FONT, size=13, bold=True)

    headers = ["키워드", "상품", "우리 순위", "우리 가격", "시장 최저가",
               "시장 평균가", "중앙값", "최저가 대비", "평균 대비",
               "가격 백분위", "판정", "비교 대상수"]
    widths = [30, 12, 10, 12, 12, 12, 12, 11, 11, 11, 14, 16]
    hr = 8
    _write_header(ws, hr, headers, widths)

    for i, res in enumerate(results):
        row = hr + 1 + i
        m = res["market"]
        vals = [
            res["keyword"], res["product"],
            res["our_rank"] if res["our_rank"] else "1000위 밖",
            res["our_price"] or "-",
            m.get("min", "-"), m.get("avg", "-"), m.get("median", "-"),
            f'{res["vs_min_pct"]:+.1f}%' if res["vs_min_pct"] is not None else "-",
            f'{res["vs_avg_pct"]:+.1f}%' if res["vs_avg_pct"] is not None else "-",
            f'하위 {res["percentile"]}%' if res["percentile"] is not None else "-",
            res["verdict"],
            f'{res["market"].get("count", 0)}개'
            + (f' (단위상이 {res["excluded"]} 제외)' if res.get("excluded") else ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = BASE_FONT
            cell.border = THIN
            if c in (4, 5, 6, 7) and isinstance(v, int):
                cell.number_format = "#,##0"
            if c >= 3:
                cell.alignment = Alignment(horizontal="center")
        vf = VERDICT_FILL.get(res["verdict"])
        if vf:
            ws.cell(row=row, column=11).fill = vf
        rank_cell = ws.cell(row=row, column=3)
        if res["our_rank"] and res["our_rank"] <= 10:
            rank_cell.font = Font(name=FONT, size=10, bold=True, color="1F7A33")
        elif not res["our_rank"]:
            rank_cell.font = Font(name=FONT, size=10, color="C00000")

    note = hr + len(results) + 2
    ws.cell(row=note, column=1,
            value="* 가격 백분위: 해당 키워드 비교 풀에서 우리 가격보다 싼 상품의 비율 (낮을수록 저렴한 편)"
            ).font = Font(name=FONT, size=9, color="808080")
    ws.cell(row=note + 1, column=1,
            value="* 단위상이 제외: 낱개/박스 등 판매 단위가 달라 보이는 상품(우리 가격과 큰 폭 차이)은 평균 왜곡 방지를 위해 가격 비교에서 자동 제외"
            ).font = Font(name=FONT, size=9, color="808080")
    ws.cell(row=note + 2, column=1,
            value="* 판정 기준·비교 범위: config.yaml의 pricing_rules에서 조정 가능"
            ).font = Font(name=FONT, size=9, color="808080")

    # ---------- 키워드별 상세 시트 ----------
    for res in results:
        wsd = wb.create_sheet(_sheet_name(res["product"] or res["keyword"]))
        wsd.sheet_view.showGridLines = False
        wsd["B1"] = f'키워드: {res["keyword"]}'
        wsd["B1"].font = Font(name=FONT, size=12, bold=True)
        band_txt = ""
        if res.get("band"):
            b = res["band"]
            hi = f'{b[1]:,.0f}' if b[1] != float("inf") else "∞"
            band_txt = f'   |   비교 범위: {b[0]:,.0f}~{hi}원 (단위상이 {res.get("excluded", 0)}개 제외)'
        wsd["B2"] = (f'우리 순위: {res["our_rank"] or "1000위 밖"}   |   '
                     f'우리 가격: {res["our_price"] or "-"}   |   판정: {res["verdict"]}{band_txt}')
        wsd["B2"].font = Font(name=FONT, size=9, color="595959")

        headers = ["순위", "판매처", "상품명", "가격", "우리 가격 대비",
                   "수량(추정)", "개당가(추정)", "상품유형", "링크"]
        widths = [7, 24, 55, 12, 13, 10, 12, 16, 40]
        _write_header(wsd, 4, headers, widths)

        our_price = res["our_price"]
        type_label = {1: "가격비교 카탈로그", 2: "일반", 3: "가격비교 매칭"}
        for i, it in enumerate(res["competitors"]):
            row = 5 + i
            diff = (f'{(it["lprice"] / our_price - 1) * 100:+.1f}%'
                    if our_price and it["lprice"] else "-")
            qty = parse_qty(it["title"])
            unit = round(it["lprice"] / qty) if qty and it["lprice"] else None
            vals = [it["rank"], mall_display(it["mallName"]), it["title"], it["lprice"],
                    diff, qty or "-", unit or "-",
                    type_label.get(it["productType"], str(it["productType"])),
                    it["link"]]
            for c, v in enumerate(vals, 1):
                cell = wsd.cell(row=row, column=c, value=v)
                cell.font = BASE_FONT
                cell.border = THIN
                if c in (4, 7):
                    cell.number_format = "#,##0"
                if c in (1, 4, 5, 6, 7, 8):
                    cell.alignment = Alignment(horizontal="center")
            if any(o["rank"] == it["rank"] for o in res["all_our_items"]):
                for c in range(1, 10):
                    wsd.cell(row=row, column=c).fill = OUR_FILL
                wsd.cell(row=row, column=2).font = BOLD_FONT

        # 우리 상품이 비교 풀 밖(순위 밖·단위 상이)이어도 전부 표기
        extra = [o for o in res["all_our_items"]
                 if not any(c["rank"] == o["rank"] for c in res["competitors"])]
        if extra:
            row = 5 + len(res["competitors"]) + 1
            wsd.cell(row=row, column=2,
                     value="※ 비교 풀 밖의 우리 상품 (참고)").font = BOLD_FONT
            for j, o in enumerate(extra):
                rr = row + 1 + j
                for c, v in enumerate([o["rank"], o["mallName"], o["title"],
                                       o["lprice"], "", "", "", "", o["link"]], 1):
                    cell = wsd.cell(row=rr, column=c, value=v)
                    cell.font = BASE_FONT
                    cell.fill = OUR_FILL
                    if c == 4:
                        cell.number_format = "#,##0"

    # ---------- 검색량 트렌드 시트 (트렌드 데이터가 있을 때만) ----------
    with_trend = [r for r in results if r.get("trend")]
    if with_trend:
        wst = wb.create_sheet("검색량트렌드", 1)
        wst.sheet_view.showGridLines = False
        wst["B1"] = "검색량 트렌드 & 기회 키워드"
        wst["B1"].font = Font(name=FONT, size=12, bold=True)
        wst["B2"] = "네이버 데이터랩 12개월 검색 추세 × 우리 순위 결합 판정 (수집 시점 기준)"
        wst["B2"].font = Font(name=FONT, size=9, color="595959")

        headers = ["키워드", "최근 3개월 추세", "변화율", "검색량 크기",
                   "피크 시즌", "우리 순위", "판정", "권장 액션"]
        widths = [28, 14, 10, 12, 10, 10, 12, 46]
        _write_header(wst, 4, headers, widths)

        DIR_TXT = {"up": "↑ 상승", "down": "↓ 하락", "flat": "→ 유지", "none": "-"}
        ACTION = {
            "기회": "상품명·태그에 키워드 반영, 블로그 콘텐츠로 노출 보완",
            "강점": "현재 순위·가격 유지, 재고 확보",
            "유지": "시즌 지남 — 현상 유지",
            "관망": "우선순위 낮음",
            "데이터부족": "검색량 매우 적음 — 더 일반적인 키워드 검토",
        }
        TAG_FILL = {"기회": PatternFill("solid", fgColor="DDEBF7"),
                    "강점": GOOD_FILL, "유지": None, "관망": None,
                    "데이터부족": None}

        def size_txt(v):
            if v is None:
                return "-"
            return "매우 큼" if v >= 1.5 else "큼" if v >= 0.8 else                    "보통" if v >= 0.4 else "작음"

        for i, r in enumerate(with_trend):
            t = r["trend"]
            row = 5 + i
            ratio = t.get("trend_ratio")
            pct = f'{(ratio - 1) * 100:+.0f}%' if ratio else "-"
            tag = t.get("tag", "관망")
            vals = [r["keyword"], DIR_TXT.get(t.get("direction"), "-"), pct,
                    size_txt(t.get("size_index")),
                    f'{t["peak_month"]}월' if t.get("peak_month") else "-",
                    r["our_rank"] if r["our_rank"] else "1000위 밖",
                    ("🎯 " if tag == "기회" else "💪 " if tag == "강점" else "") + tag,
                    ACTION.get(tag, "")]
            for c, v in enumerate(vals, 1):
                cell = wst.cell(row=row, column=c, value=v)
                cell.font = BASE_FONT
                cell.border = THIN
                if c in (2, 3, 4, 5, 6, 7):
                    cell.alignment = Alignment(horizontal="center")
            f = TAG_FILL.get(tag)
            if f:
                for c in range(1, 9):
                    wst.cell(row=row, column=c).fill = f
            if tag == "기회":
                wst.cell(row=row, column=7).font = Font(name=FONT, size=10,
                                                        bold=True, color="1C5CAB")

        note = 5 + len(with_trend) + 1
        wst.cell(row=note, column=1,
                 value="* 검색량은 네이버 데이터랩 상대값(기간 내 최고=100) 기반 — 정확한 검색 횟수가 아닌 흐름·비교용 지표"
                 ).font = Font(name=FONT, size=9, color="808080")
        wst.cell(row=note + 1, column=1,
                 value="* 기회 = 검색량 활발(상승 또는 규모 큼) + 우리 순위 100위 밖 → 노출 개선 시 매출 기대가 큰 키워드"
                 ).font = Font(name=FONT, size=9, color="808080")

        # 기회 키워드별 개선 가이드
        opps = [r for r in with_trend if r["trend"].get("tag") == "기회"]
        if opps:
            row = note + 3
            head = wst.cell(row=row, column=1, value="■ 기회 키워드 개선 가이드")
            head.font = Font(name=FONT, size=11, bold=True)
            row += 1
            for r in opps:
                kw = r["keyword"]
                nospace = kw.replace(" ", "")
                peak = r["trend"].get("peak_month")
                rank_txt = f'{r["our_rank"]}위' if r["our_rank"] else "1000위 밖"
                steps = [
                    f'🎯 "{kw}"  (현재 {rank_txt})',
                    f'  1. 상품명에 키워드 넣기(효과 가장 큼): 스마트스토어센터 → 상품관리에서 상품명에 "{kw}" 표현을 자연스럽게 포함',
                    f'  2. 태그 10개 채우기: 상품 수정 → 검색설정 → 태그에 "{nospace}", "{kw}" 등 붙여쓰기·띄어쓰기·연관 표현 추가',
                    '  3. 카테고리·속성 점검: 키워드와 맞는 카테고리인지 확인, 속성 정보 빈칸 없이',
                    f'  4. 블로그 콘텐츠 발행: "{kw}" 제목의 사용기·추천 글 발행 (링커 블로그잇 대행 가능)',
                ]
                if peak:
                    steps.append(f'  5. 타이밍: 피크 시즌 {peak}월 기준 1~2개월 전까지 작업·재고 준비 완료')
                steps.append('  6. 적용 3~7일 후 다시 수집해서 순위 변화 확인')
                for line in steps:
                    c = wst.cell(row=row, column=1, value=line)
                    bold = line.startswith("🎯")
                    c.font = Font(name=FONT, size=10, bold=bold,
                                  color="1C5CAB" if bold else "404040")
                    row += 1
                row += 1

    wb.save(out_path)
    return out_path
